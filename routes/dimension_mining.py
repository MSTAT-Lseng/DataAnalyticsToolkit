"""维度挖掘 API — 基于关键词配置对文本列进行多维度标注与评分。"""

import json as _json

import pandas as pd

from flask import Blueprint, current_app, jsonify, request

from utils.dimension_mining import mine_dimensions
from utils.file_helpers import ALLOWED_TABLE_EXTENSIONS, uploaded_dataframe, dataframe_preview, send_excel

dimension_mining_bp = Blueprint("dimension_mining", __name__, url_prefix="/api/dimension-mining")


def _validate_dimensions(dimensions: list) -> str | None:
    """验证维度配置结构，返回错误信息或 None。"""
    if not isinstance(dimensions, list) or len(dimensions) == 0:
        return "请配置至少一个维度"

    for i, dim in enumerate(dimensions):
        if not dim.get("name", "").strip():
            return f"第 {i + 1} 个维度缺少名称"
        keywords = dim.get("keywords", [])
        if not keywords or not isinstance(keywords, list):
            return f"维度「{dim['name']}」至少需要一个关键词"
        for j, kw in enumerate(keywords):
            if not kw.get("pattern", "").strip():
                return f"维度「{dim['name']}」的第 {j + 1} 个关键词缺少匹配模式"

    return None


@dimension_mining_bp.route("/preview", methods=["POST"])
def api_dimension_mining_preview():
    """接收表格文件，返回列名和预览数据。"""
    try:
        file = request.files.get("file")
        with uploaded_dataframe(file) as (df, _suffix, _tmp_path):
            preview = dataframe_preview(df)
            current_app.logger.info(
                "Dimension mining preview: %d columns, %d preview rows (total %d rows)",
                len(preview["columns"]), len(preview["rows"]), preview["total_rows"],
            )
            return jsonify({"success": True, **preview})
    except ValueError as exc:
        return jsonify({"success": False, "error": str(exc)}), 400
    except Exception as exc:
        current_app.logger.exception("Dimension mining preview API error")
        return jsonify({"success": False, "error": str(exc)}), 500


@dimension_mining_bp.route("/analyze", methods=["POST"])
def api_dimension_mining_analyze():
    """接收表格文件 + 列名 + 维度配置，进行维度挖掘分析。"""
    try:
        file = request.files.get("file")
        if not file or file.filename == "":
            return jsonify({"success": False, "error": "请上传表格文件"}), 400

        column = (request.form.get("column") or "").strip()
        if not column:
            return jsonify({"success": False, "error": "请选择要分析的列"}), 400

        dimensions_raw = (request.form.get("dimensions") or "").strip()
        if not dimensions_raw:
            return jsonify({"success": False, "error": "请配置至少一个维度"}), 400

        try:
            dimensions = _json.loads(dimensions_raw)
        except _json.JSONDecodeError as exc:
            return jsonify({
                "success": False,
                "error": f"维度配置 JSON 格式错误：{exc}"
            }), 400

        error = _validate_dimensions(dimensions)
        if error:
            return jsonify({"success": False, "error": error}), 400

        with uploaded_dataframe(file) as (df, _suffix, _tmp_path):
            if column not in df.columns:
                available = [str(c) for c in df.columns.tolist()]
                return jsonify({
                    "success": False,
                    "error": f"列 '{column}' 不存在。可用列：{', '.join(available)}"
                }), 400

            col_data = df[column].dropna().astype(str).tolist()

            if len(col_data) == 0:
                return jsonify({
                    "success": False,
                    "error": f"列 '{column}' 中没有有效文本"
                }), 400

            result = mine_dimensions(col_data, dimensions)

            current_app.logger.info(
                "Dimension mining done: column='%s', %d rows, %d dimensions",
                column, result["total_rows"], result["dimension_count"],
            )

            return jsonify({
                "success": True,
                "result": result,
                "source_column": column,
                "source_file": file.filename,
            })
    except ValueError as exc:
        return jsonify({"success": False, "error": str(exc)}), 400
    except Exception as exc:
        current_app.logger.exception("Dimension mining analyze API error")
        return jsonify({"success": False, "error": str(exc)}), 500


@dimension_mining_bp.route("/export", methods=["POST"])
def api_dimension_mining_export():
    """接收分析结果 JSON，生成 Excel 文件并返回下载。"""
    try:
        import openpyxl

        data = request.get_json(force=True)
        result = data.get("result") or {}
        source_info = data.get("source_info", "")

        dimensions = result.get("dimensions", [])
        row_results = result.get("row_results", [])

        if not dimensions or not row_results:
            return jsonify({"success": False, "error": "没有可导出的数据"}), 400

        wb = openpyxl.Workbook()

        # ---- Sheet 1: 维度汇总 ----
        ws_summary = wb.active
        ws_summary.title = "维度汇总"
        ws_summary.append(["维度名称", "整体得分", "匹配行数", "匹配率", "关键词匹配详情"])
        for dim in dimensions:
            kw_detail = "; ".join(
                f"{kw}: {cnt}次" for kw, cnt in dim.get("keywords_matched", {}).items()
            )
            ws_summary.append([
                dim["name"],
                dim["overall_score"],
                dim["match_count"],
                f"{dim['match_rate'] * 100:.1f}%",
                kw_detail,
            ])
        ws_summary.append([])
        ws_summary.append(["数据来源", source_info])
        ws_summary.append(["总行数", result.get("total_rows", len(row_results))])
        ws_summary.append(["维度数量", result.get("dimension_count", len(dimensions))])

        ws_summary.column_dimensions["A"].width = 16
        ws_summary.column_dimensions["B"].width = 12
        ws_summary.column_dimensions["C"].width = 12
        ws_summary.column_dimensions["D"].width = 10
        ws_summary.column_dimensions["E"].width = 50

        # ---- Sheet 2: 逐行详细结果 ----
        ws_detail = wb.create_sheet(title="逐行分析结果")
        dim_names = [d["name"] for d in dimensions]
        headers = ["#", "文本内容"] + [f"{name}（得分）" for name in dim_names] + \
                  [f"{name}（匹配词）" for name in dim_names]
        ws_detail.append(headers)

        for i, row in enumerate(row_results):
            row_data = [i + 1, row.get("text", "")]
            scores = row.get("scores", {})
            matches = row.get("matches", {})
            for name in dim_names:
                row_data.append(scores.get(name, 0))
            for name in dim_names:
                row_data.append(", ".join(matches.get(name, [])))
            ws_detail.append(row_data)

        ws_detail.column_dimensions["A"].width = 6
        ws_detail.column_dimensions["B"].width = 50
        for j, name in enumerate(dim_names):
            col_letter = openpyxl.utils.get_column_letter(3 + j)
            ws_detail.column_dimensions[col_letter].width = 14
            col_letter2 = openpyxl.utils.get_column_letter(3 + len(dim_names) + j)
            ws_detail.column_dimensions[col_letter2].width = 24

        return send_excel(wb, "维度挖掘结果.xlsx")
    except Exception as exc:
        current_app.logger.exception("Dimension mining export API error")
        return jsonify({"success": False, "error": str(exc)}), 500
